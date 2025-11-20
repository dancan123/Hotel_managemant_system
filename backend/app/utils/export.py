"""Export utilities for Excel and PDF generation"""
import os
import json
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

def export_to_excel(data, filename, title="Report"):
    """Export data to Excel file"""
    if not OPENPYXL_AVAILABLE:
        return None, "openpyxl not installed"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add title
    ws['A1'] = title
    title_font = Font(bold=True, size=14)
    ws['A1'].font = title_font
    
    # Add timestamp
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Add headers
    if data and len(data) > 0:
        headers = data[0].keys() if isinstance(data[0], dict) else list(range(len(data[0])))
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row_idx, row_data in enumerate(data, start=5):
            if isinstance(row_data, dict):
                for col_idx, (key, value) in enumerate(row_data.items(), start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    try:
        wb.save(filename)
        return filename, "Success"
    except Exception as e:
        return None, str(e)

def export_to_pdf(data, filename, title="Report"):
    """Export data to PDF file"""
    if not REPORTLAB_AVAILABLE:
        return None, "reportlab not installed"
    
    try:
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Add title
        title_style = styles['Heading1']
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.3*inch))
        
        # Convert data to table format
        if data:
            if isinstance(data[0], dict):
                headers = list(data[0].keys())
                table_data = [headers]
                for row in data:
                    table_data.append([str(row.get(key, '')) for key in headers])
            else:
                table_data = [list(row) for row in data]
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            elements.append(table)
        
        # Add generation timestamp
        elements.append(Spacer(1, 0.3*inch))
        timestamp = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        elements.append(Paragraph(timestamp, styles['Normal']))
        
        doc.build(elements)
        return filename, "Success"
    except Exception as e:
        return None, str(e)

def export_summary_report(sales_data, employees_data, period, export_format='excel'):
    """Export comprehensive summary report"""
    filename = f"hotel_report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{'xlsx' if export_format == 'excel' else 'pdf'}"
    
    if export_format == 'excel':
        return export_to_excel(sales_data, filename, f"Sales Report - {period}")
    else:
        return export_to_pdf(sales_data, filename, f"Sales Report - {period}")
